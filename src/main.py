import os
import datetime
import requests
import openpyxl
from dotenv import load_dotenv

# Jira API URL and API Key
JIRA_API_URL = "https://safeway.atlassian.net/rest/api/latest/search?jql=filter="

def get_jira_data(filter_id):
    auth = (os.getenv('JIRA_USER'), os.getenv('JIRA_API_TOKEN'))
    url = f'{JIRA_API_URL}{filter_id}'
    
    data = []
    startAt = 0
    maxResults = 100
    total = 1
    
    while startAt <= total:
        params = {
            'startAt': startAt,
            'maxResults': maxResults
        }
        response = requests.get(url, auth=auth, params=params)
        res_json = response.json()
        data.extend(res_json['issues'])
        startAt += maxResults
        total = res_json['total']
    
    if response.status_code == 200:
        return data
    else:
        print(f'Failed to fetch Jira data for filter ID {filter_id}')
        return []

def insert_epics(ws, data):
    ws.delete_rows(1, ws.max_row)
    headers = ['Issue Key', 'Summary', 'Assignee', 'Status', 'Project Name']
    ws.append(headers)

    for issue in data:
        issue_key = issue['key']
        summary = issue['fields']['summary']
        assignee = issue['fields']['assignee']['displayName'] if issue['fields']['assignee'] else 'Unassigned'
        status = issue['fields']['status']['name']
        project_name = issue['fields']['project']['name']
        ws.append([issue_key, summary, assignee, status, project_name])

def insert_stories(ws, data):
    ws.delete_rows(1, ws.max_row)
    headers = ['Issue Key', 'Summary', 'Assignee', 'Status', 'Story Point', 'Sprint', 'Parent Key', 'Parent', 'Project']
    ws.append(headers)

    for issue in data:
        issue_key = issue['key']
        summary = issue['fields']['summary']
        assignee = issue['fields']['assignee']['displayName'] if issue['fields']['assignee'] else 'Unassigned'
        status = issue['fields']['status']['name']
        story_point = issue['fields'].get('customfield_10002', '')  # Assuming story points are stored in this custom field
        sprint = issue['fields'].get('sprint', {}).get('name', '') if 'sprint' in issue['fields'] else ''
        parent_key = issue['fields'].get('parent', {}).get('key', '')
        parent_summary = issue['fields'].get('parent', {}).get('fields', {}).get('summary', '')
        project_name = issue['fields']['project']['name']
        ws.append([issue_key, summary, assignee, status, story_point, sprint, parent_key, parent_summary, project_name])

def insert_bugs(ws, data):
    ws.delete_rows(1, ws.max_row)
    headers = ['Issue Key', 'Summary', 'Assignee', 'Status', 'Priority', 'Project']
    ws.append(headers)

    for issue in data:
        issue_key = issue['key']
        summary = issue['fields']['summary']
        assignee = issue['fields']['assignee']['displayName'] if issue['fields']['assignee'] else 'Unassigned'
        status = issue['fields']['status']['name']
        project_name = issue['fields']['project']['name']
        priority = issue['fields'].get('priority', {}).get('name', 'None')
        ws.append([issue_key, summary, assignee, status, priority, project_name])

def update_timestamp(wb):
    ws = wb["Summary"]
    ws["H1"] = datetime.datetime.now()

def write_to_excel(data, wb, worksheet_name, type):
    ws = wb[worksheet_name]
    
    if type == "epics":
        insert_epics(ws, data)
    elif type == "stories":
        insert_stories(ws, data)
    elif type == "bugs":
        insert_bugs(ws, data)
    print(f'Data written to {worksheet_name}')

def main():
    # Load environment variables from .env file
    load_dotenv()

    # Check for required environment variables
    required_vars = ['JIRA_URL', 'JIRA_USER', 'JIRA_API_TOKEN']
    missing_vars = [var for var in required_vars if not os.getenv(var)]
    
    if missing_vars:
        print(f"Missing required environment variables: {', '.join(missing_vars)}. Exiting...")
        return

    file_path = os.getenv('EXCEL_FILE_PATH')
    if not file_path:
        print("Excel file path not specified in environment variables. Exiting...")
        return

    wb = openpyxl.load_workbook(filename=file_path)

    # Report Epics
    filter_id = os.getenv('EPIC_FILTER_ID')
    worksheet_name = 'Epics Raw'
    epics_data = get_jira_data(filter_id)
    if epics_data:
        write_to_excel(epics_data, wb, worksheet_name, "epics")

    # Report Stories - All
    filter_id = os.getenv('ALL_STORIES_FILTER_ID')
    worksheet_name = 'Stories Raw'
    stories_data = get_jira_data(filter_id)
    if stories_data:
        write_to_excel(stories_data, wb, worksheet_name, "stories")

    # Report Stories - Active
    filter_id = os.getenv('ACTIVE_STORIES_FILTER_ID')
    worksheet_name = 'Stories Raw (Active Sprint)'
    active_stories_data = get_jira_data(filter_id)
    if active_stories_data:
        write_to_excel(active_stories_data, wb, worksheet_name, "stories")

    # Report bugs
    filter_id = os.getenv('BUGS_FILTER_ID')
    worksheet_name = 'Bugs Raw'
    bugs_data = get_jira_data(filter_id)
    if bugs_data:
        write_to_excel(bugs_data, wb, worksheet_name, "bugs")

    update_timestamp(wb)
    wb.save(filename=file_path)
    print(f"Data exported successfully to {file_path}")

if __name__ == "__main__":
    main()